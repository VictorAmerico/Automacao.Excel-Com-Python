import openpyxl

#criar uma planilha(book)

book = openpyxl.Workbook();

#Como visualizar páginas existentes

print(book.sheetnames)

#Como criar uma página

book.create_sheet('Frutas')

#Como selecionar uma página

frutas_page = book['Frutas']
frutas_page.append(['Produto','Quantidade','Preço'])
frutas_page.append(['Banana','5','R$3,90'])
frutas_page.append(['Fruta 2','2','R$3,90'])
frutas_page.append(['Fruta 3','10','R$3,90'])
frutas_page.append(['Fruta 4','4','R$3,90'])

#Sakvar a planilha

book.save('Planilha de compras.xlsx')